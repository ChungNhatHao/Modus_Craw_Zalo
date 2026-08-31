from __future__ import annotations

import hashlib
import json
import os
import re
import uuid
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote

import openpyxl

from .models import CleanMessage, ImageOcrResult, MediaAsset, OrderDecision
from .storage import safe_slug


DRIVE_API = "https://www.googleapis.com/drive/v3"
DRIVE_UPLOAD_API = "https://www.googleapis.com/upload/drive/v3"
SHEETS_API = "https://sheets.googleapis.com/v4"
GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
GOOGLE_FOLDER_MIME = "application/vnd.google-apps.folder"
GOOGLE_SCOPES = (
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets",
)
SHEET_MARKER_KEY = "zaloOrderCrawler"
SHEET_MARKER_VALUE = "daily-text-v1"
BRANCH_CONFIG_MARKER_VALUE = "branch-config-v1"
BRANCH_CONFIG_NAME = "Cấu hình chi nhánh"
BRANCH_CONFIG_TAB = "Chi nhánh"
BRANCH_CONFIG_HEADERS = ("Tên nhận diện", "Chi nhánh chuẩn")
DEFAULT_BRANCH_MAPPINGS = (
    ("S6", "Chi nhánh Phạm Văn Đồng"),
    ("Tân Phú", "Chi nhánh Tân Phú"),
    ("Sườn Thảo Điền", "Chi nhánh Thảo Điền"),
)
IMAGE_MARKER_KEY = "zaloCrawlerImageKey"
OCR_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
OCR_MARKER_VALUE = "order-image-ocr-v1"
OCR_SHEET_TITLE = "OCR"
OCR_HEADERS = (
    "Ngày",
    "Nhóm",
    "Mã tin nhắn",
    "Mã khách hàng",
    "Tên khách hàng",
    "Tên hàng",
    "Đơn vị",
    "Số lượng",
)
LEGACY_SHEET_HEADERS = (
    "Ngày",
    "Nhóm",
    "Mã tin nhắn",
    "STT",
    "Người gửi",
    "Thời gian",
    "Chiều tin nhắn",
    "Nội dung",
    "Loại tin",
)
SHEET_HEADERS = (*LEGACY_SHEET_HEADERS, "Chi nhánh")
_MEDIA_PLACEHOLDERS = {"[hình ảnh]", "[image]"}
_DRIVE_ID_PATTERN = re.compile(r"^[0-9A-Za-z_-]+$")


class GoogleDriveOutputError(RuntimeError):
    """Raised when daily output cannot be written safely to Google Drive."""


@dataclass(frozen=True)
class DriveResource:
    id: str
    name: str
    url: str


@dataclass(frozen=True)
class BranchConfig:
    resource: DriveResource
    mappings: dict[str, str]


class GoogleDriveOutputPublisher:
    def __init__(self, session: Any, parent_folder_id: str) -> None:
        if not _DRIVE_ID_PATTERN.fullmatch(parent_folder_id):
            raise ValueError("GOOGLE_DRIVE_PARENT_FOLDER_ID không hợp lệ.")
        self.session = session
        self.parent_folder_id = parent_folder_id

    @classmethod
    def from_default_credentials(
        cls, parent_folder_id: str, *, project_dir: Path | None = None
    ) -> "GoogleDriveOutputPublisher":
        try:
            import google.auth
            from google.auth.transport.requests import Request
            from google.auth.transport.requests import AuthorizedSession

            client_secret_value = os.environ.get(
                "GOOGLE_OAUTH_CLIENT_SECRET_FILE", ""
            ).strip()
            if client_secret_value:
                from google.oauth2.credentials import Credentials
                from google_auth_oauthlib.flow import InstalledAppFlow

                base_dir = (project_dir or Path.cwd()).resolve()
                client_secret = cls._configured_path(client_secret_value, base_dir)
                token_value = os.environ.get(
                    "GOOGLE_OAUTH_TOKEN_FILE", ".google-drive-token.json"
                ).strip()
                token_file = cls._configured_path(token_value, base_dir)
                credentials = None
                if token_file.is_file():
                    try:
                        credentials = Credentials.from_authorized_user_file(
                            str(token_file), GOOGLE_SCOPES
                        )
                    except (OSError, ValueError):
                        credentials = None
                if credentials and credentials.expired and credentials.refresh_token:
                    try:
                        credentials.refresh(Request())
                    except Exception:
                        credentials = None
                if not credentials or not credentials.valid:
                    flow = InstalledAppFlow.from_client_secrets_file(
                        str(client_secret), GOOGLE_SCOPES
                    )
                    client_config = json.loads(
                        client_secret.read_text(encoding="utf-8")
                    )
                    if "web" in client_config:
                        redirect_port = cls._oauth_redirect_port()
                        credentials = flow.run_local_server(
                            host="127.0.0.1",
                            bind_addr="127.0.0.1",
                            port=redirect_port,
                            timeout_seconds=300,
                            authorization_prompt_message=(
                                "Hãy mở URL này để cấp quyền Google Drive: {url}"
                            ),
                            success_message=(
                                "Đã xác thực Google thành công. Bạn có thể đóng tab này."
                            ),
                        )
                    else:
                        credentials = flow.run_local_server(
                            port=0,
                            timeout_seconds=300,
                            success_message=(
                                "Đã xác thực Google thành công. Bạn có thể đóng tab này."
                            ),
                        )
                token_file.parent.mkdir(parents=True, exist_ok=True)
                token_file.write_text(credentials.to_json(), encoding="utf-8")
                try:
                    token_file.chmod(0o600)
                except OSError:
                    pass
            else:
                credentials, _ = google.auth.default(scopes=GOOGLE_SCOPES)
        except Exception as exc:
            raise GoogleDriveOutputError(
                "Chưa xác thực được Google. Hãy cấu hình "
                "GOOGLE_OAUTH_CLIENT_SECRET_FILE để đăng nhập tài khoản Google, "
                "hoặc GOOGLE_APPLICATION_CREDENTIALS/Application Default Credentials."
            ) from exc
        return cls(AuthorizedSession(credentials), parent_folder_id)

    def verify_destination(self) -> DriveResource:
        payload = self._request_json(
            "GET",
            f"{DRIVE_API}/files/{quote(self.parent_folder_id, safe='')}",
            params={
                "supportsAllDrives": "true",
                "fields": (
                    "id,name,mimeType,webViewLink,trashed,"
                    "capabilities(canAddChildren)"
                ),
            },
        )
        if payload.get("mimeType") != GOOGLE_FOLDER_MIME or payload.get("trashed"):
            raise GoogleDriveOutputError(
                "GOOGLE_DRIVE_PARENT_FOLDER_ID không trỏ tới một folder đang hoạt động."
            )
        capabilities = payload.get("capabilities") or {}
        if capabilities.get("canAddChildren") is not True:
            raise GoogleDriveOutputError(
                "Tài khoản Google hiện tại không có quyền tạo file trong folder đích."
            )
        return self._resource(payload)

    def publish(
        self,
        *,
        run_dir: Path,
        group_name: str,
        target_date: date,
        messages: Iterable[CleanMessage],
        decisions: Iterable[OrderDecision] = (),
        ocr_results: Iterable[ImageOcrResult] = (),
    ) -> dict[str, Any]:
        run_dir = run_dir.resolve()
        message_list = list(messages)
        text_rows = self._text_rows(
            group_name,
            target_date,
            message_list,
            decisions,
        )
        image_assets = self._message_images(run_dir, message_list)
        daily_name = target_date.strftime("%d-%m-%Y")
        output: dict[str, Any] = {}

        if text_rows:
            sheet, sheet_title, row_count = self._ensure_daily_sheet(daily_name)
            rows_added = self._append_unique_text_rows(
                sheet.id, sheet_title, row_count, text_rows
            )
            output["sheet"] = {
                "id": sheet.id,
                "name": sheet.name,
                "url": sheet.url,
                "rows_added": rows_added,
            }

        if image_assets:
            folder = self._ensure_image_folder(f"{daily_name}_image")
            uploaded = 0
            existing = 0
            for message_id, media, media_path in image_assets:
                was_uploaded = self._upload_unique_image(
                    folder.id,
                    group_name=group_name,
                    message_id=message_id,
                    media=media,
                    media_path=media_path,
                )
                uploaded += int(was_uploaded)
                existing += int(not was_uploaded)
            output["image_folder"] = {
                "id": folder.id,
                "name": folder.name,
                "url": folder.url,
                "images_uploaded": uploaded,
                "images_existing": existing,
            }

        ocr_rows = self._ocr_rows(group_name, target_date, ocr_results)
        if ocr_rows:
            workbook_resource = self._ensure_ocr_workbook(f"{daily_name}_OCR.xlsx")
            rows_added = self._append_unique_ocr_rows(workbook_resource.id, ocr_rows)
            output["ocr_workbook"] = {
                "id": workbook_resource.id,
                "name": workbook_resource.name,
                "url": workbook_resource.url,
                "rows_added": rows_added,
            }

        return output

    @staticmethod
    def _text_rows(
        group_name: str,
        target_date: date,
        messages: Iterable[CleanMessage],
        decisions: Iterable[OrderDecision] = (),
    ) -> list[list[Any]]:
        display_date = target_date.strftime("%d-%m-%Y")
        branch_by_message_id = {
            decision.message_id: decision.branch_name or ""
            for decision in decisions
            if decision.is_order
        }
        rows: list[list[Any]] = []
        for message in messages:
            content = message.content.strip()
            if not content or content.casefold() in _MEDIA_PLACEHOLDERS:
                continue
            rows.append(
                [
                    display_date,
                    group_name,
                    message.message_id,
                    message.sequence,
                    message.sender or "",
                    message.timestamp_text or "",
                    message.direction,
                    content,
                    message.message_type,
                    branch_by_message_id.get(message.message_id, ""),
                ]
            )
        return rows

    @classmethod
    def _message_images(
        cls,
        run_dir: Path,
        messages: Iterable[CleanMessage],
    ) -> list[tuple[str, MediaAsset, Path]]:
        result: list[tuple[str, MediaAsset, Path]] = []
        seen: set[tuple[str, str]] = set()
        for message in messages:
            for media in message.media:
                if media.role != "message_image" or not media.mime_type.startswith("image/"):
                    continue
                media_path = cls._resolve_media_path(run_dir, media.path)
                identity = (media.sha256, str(media_path))
                if identity in seen:
                    continue
                seen.add(identity)
                result.append((message.message_id, media, media_path))
        return result

    @staticmethod
    def _resolve_media_path(run_dir: Path, media_path: str) -> Path:
        raw = Path(media_path)
        if not media_path or raw.is_absolute():
            raise GoogleDriveOutputError("Đường dẫn ảnh output không hợp lệ.")
        candidate = (run_dir / raw).resolve()
        try:
            candidate.relative_to(run_dir)
        except ValueError as exc:
            raise GoogleDriveOutputError(
                "Ảnh output nằm ngoài thư mục của lượt crawl."
            ) from exc
        if not candidate.is_file():
            raise GoogleDriveOutputError(f"Không tìm thấy ảnh output: {media_path}")
        return candidate

    @staticmethod
    def _ocr_rows(
        group_name: str,
        target_date: date,
        ocr_results: Iterable[ImageOcrResult],
    ) -> list[list[Any]]:
        display_date = target_date.strftime("%d-%m-%Y")
        rows: list[list[Any]] = []
        for result in ocr_results:
            if not result.applicable:
                continue
            for item in result.items:
                if not item.product_name.strip():
                    continue
                rows.append(
                    [
                        display_date,
                        group_name,
                        result.message_id,
                        item.customer_code,
                        item.customer_name,
                        item.product_name,
                        item.unit,
                        item.quantity if item.quantity is not None else "",
                    ]
                )
        return rows

    def _ensure_ocr_workbook(self, name: str) -> DriveResource:
        marker_query = (
            f"appProperties has {{ key='{SHEET_MARKER_KEY}' and "
            f"value='{OCR_MARKER_VALUE}' }}"
        )
        resource = self._find_resource(
            self.parent_folder_id, name, OCR_MIME, extra_query=marker_query
        )
        if resource is not None:
            return resource
        empty_workbook = self._build_workbook(OCR_HEADERS, [])
        return self._create_binary_resource(
            name,
            OCR_MIME,
            self.parent_folder_id,
            empty_workbook,
            app_properties={SHEET_MARKER_KEY: OCR_MARKER_VALUE},
        )

    def _append_unique_ocr_rows(self, file_id: str, rows: list[list[Any]]) -> int:
        existing_bytes = self._download_file(file_id)
        workbook = openpyxl.load_workbook(BytesIO(existing_bytes))
        sheet = workbook.active
        existing_keys: set[tuple[str, str, str]] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 6:
                existing_keys.add(
                    (str(row[1] or ""), str(row[2] or ""), str(row[5] or ""))
                )
        pending = [
            row
            for row in rows
            if (str(row[1]), str(row[2]), str(row[5])) not in existing_keys
        ]
        if not pending:
            return 0
        for row in pending:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        self._update_file_content(file_id, OCR_MIME, buffer.getvalue())
        return len(pending)

    @staticmethod
    def _build_workbook(headers: tuple[str, ...], rows: list[list[Any]]) -> bytes:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = OCR_SHEET_TITLE
        sheet.append(list(headers))
        for row in rows:
            sheet.append(row)
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _create_binary_resource(
        self,
        name: str,
        mime_type: str,
        parent_id: str,
        data: bytes,
        *,
        app_properties: dict[str, str] | None = None,
    ) -> DriveResource:
        metadata: dict[str, Any] = {"name": name, "parents": [parent_id]}
        if app_properties:
            metadata["appProperties"] = app_properties
        boundary = f"codex-{uuid.uuid4().hex}"
        metadata_bytes = json.dumps(metadata, ensure_ascii=False).encode("utf-8")
        body = b"\r\n".join(
            [
                f"--{boundary}".encode(),
                b"Content-Type: application/json; charset=UTF-8",
                b"",
                metadata_bytes,
                f"--{boundary}".encode(),
                f"Content-Type: {mime_type}".encode(),
                b"",
                data,
                f"--{boundary}--".encode(),
                b"",
            ]
        )
        payload = self._request_json(
            "POST",
            f"{DRIVE_UPLOAD_API}/files",
            expected=(200, 201),
            params={
                "uploadType": "multipart",
                "supportsAllDrives": "true",
                "fields": "id",
            },
            headers={"Content-Type": f"multipart/related; boundary={boundary}"},
            data=body,
        )
        resource_id = str(payload.get("id") or "")
        if not resource_id:
            raise GoogleDriveOutputError("Google Drive không trả id tài nguyên mới.")
        verified = self._request_json(
            "GET",
            f"{DRIVE_API}/files/{quote(resource_id, safe='')}",
            params={
                "supportsAllDrives": "true",
                "fields": "id,name,mimeType,webViewLink",
            },
        )
        return self._resource(verified)

    def _download_file(self, file_id: str) -> bytes:
        try:
            response = self.session.request(
                "GET",
                f"{DRIVE_API}/files/{quote(file_id, safe='')}",
                params={"alt": "media", "supportsAllDrives": "true"},
                timeout=120,
            )
        except Exception as exc:
            raise GoogleDriveOutputError(
                f"Không kết nối được Google API: {exc}"
            ) from exc
        if response.status_code != 200:
            raise GoogleDriveOutputError(
                f"Google API trả HTTP {response.status_code} khi tải file {file_id}."
            )
        return response.content

    def _update_file_content(self, file_id: str, mime_type: str, data: bytes) -> None:
        self._request_json(
            "PATCH",
            f"{DRIVE_UPLOAD_API}/files/{quote(file_id, safe='')}",
            params={
                "uploadType": "media",
                "supportsAllDrives": "true",
                "fields": "id,name",
            },
            headers={"Content-Type": mime_type},
            data=data,
        )

    def ensure_branch_config(self) -> BranchConfig:
        marker_query = (
            f"appProperties has {{ key='{SHEET_MARKER_KEY}' and "
            f"value='{BRANCH_CONFIG_MARKER_VALUE}' }}"
        )
        resource = self._find_resource(
            self.parent_folder_id,
            BRANCH_CONFIG_NAME,
            GOOGLE_SHEET_MIME,
            extra_query=marker_query,
        )
        created = resource is None
        if resource is None:
            resource = self._create_resource(
                BRANCH_CONFIG_NAME,
                GOOGLE_SHEET_MIME,
                self.parent_folder_id,
                app_properties={SHEET_MARKER_KEY: BRANCH_CONFIG_MARKER_VALUE},
            )

        sheet_id, sheet_title, row_count = self._first_sheet(resource.id)
        if created and sheet_title != BRANCH_CONFIG_TAB:
            self._rename_sheet(resource.id, sheet_id, BRANCH_CONFIG_TAB)
            sheet_title = BRANCH_CONFIG_TAB

        header_range = f"{self._a1_sheet(sheet_title)}!A1:B1"
        header_values = self._get_values(resource.id, header_range)
        if not header_values:
            seed_rows = [list(BRANCH_CONFIG_HEADERS), *map(list, DEFAULT_BRANCH_MAPPINGS)]
            seed_range = f"{self._a1_sheet(sheet_title)}!A1:B{len(seed_rows)}"
            self._request_json(
                "PUT",
                f"{SHEETS_API}/spreadsheets/{quote(resource.id, safe='')}/values/"
                f"{quote(seed_range, safe='')}",
                params={"valueInputOption": "RAW"},
                json={"majorDimension": "ROWS", "values": seed_rows},
            )
            self._format_branch_config(resource.id, sheet_id)
        else:
            existing_header = [str(value).strip() for value in header_values[0]]
            if existing_header != list(BRANCH_CONFIG_HEADERS):
                raise GoogleDriveOutputError(
                    f"Sheet cấu hình {resource.id} phải có header "
                    f"{', '.join(BRANCH_CONFIG_HEADERS)}."
                )

        max_row = min(max(row_count, len(DEFAULT_BRANCH_MAPPINGS) + 1), 10_000)
        data_range = f"{self._a1_sheet(sheet_title)}!A2:B{max_row}"
        mappings = self._parse_branch_mappings(
            self._get_values(resource.id, data_range)
        )
        return BranchConfig(resource=resource, mappings=mappings)

    def _format_branch_config(self, spreadsheet_id: str, sheet_id: int) -> None:
        self._sheets_batch_update(
            spreadsheet_id,
            [
                {
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": sheet_id,
                            "gridProperties": {"frozenRowCount": 1},
                        },
                        "fields": "gridProperties.frozenRowCount",
                    }
                },
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 0,
                            "endRowIndex": 1,
                            "startColumnIndex": 0,
                            "endColumnIndex": 2,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColorStyle": {
                                    "rgbColor": {
                                        "red": 0.9,
                                        "green": 0.9,
                                        "blue": 0.9,
                                    }
                                },
                                "textFormat": {"bold": True},
                            }
                        },
                        "fields": (
                            "userEnteredFormat(backgroundColorStyle,textFormat)"
                        ),
                    }
                },
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": sheet_id,
                            "dimension": "COLUMNS",
                            "startIndex": 0,
                            "endIndex": 1,
                        },
                        "properties": {"pixelSize": 180},
                        "fields": "pixelSize",
                    }
                },
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": sheet_id,
                            "dimension": "COLUMNS",
                            "startIndex": 1,
                            "endIndex": 2,
                        },
                        "properties": {"pixelSize": 260},
                        "fields": "pixelSize",
                    }
                },
            ],
        )

    @classmethod
    def _parse_branch_mappings(cls, rows: Iterable[list[Any]]) -> dict[str, str]:
        mappings: dict[str, str] = {}
        canonical_by_key: dict[str, str] = {}
        display_aliases: dict[str, str] = {}
        for row_number, row in enumerate(rows, start=2):
            alias = str(row[0]).strip() if row else ""
            canonical = str(row[1]).strip() if len(row) >= 2 else ""
            if not alias and not canonical:
                continue
            if not alias or not canonical:
                raise GoogleDriveOutputError(
                    f"Cấu hình chi nhánh dòng {row_number} phải có đủ hai cột."
                )
            key = cls._normalise_branch_key(alias)
            if key in canonical_by_key and canonical_by_key[key] != canonical:
                raise GoogleDriveOutputError(
                    f"Tên nhận diện {display_aliases[key]!r} bị ánh xạ tới nhiều "
                    "chi nhánh khác nhau."
                )
            mappings[alias] = canonical
            canonical_by_key[key] = canonical
            display_aliases[key] = alias
        if not mappings:
            raise GoogleDriveOutputError(
                "Google Sheet cấu hình chi nhánh chưa có ánh xạ nào."
            )
        return mappings

    @staticmethod
    def _normalise_branch_key(value: str) -> str:
        return " ".join(value.casefold().split())

    def _ensure_daily_sheet(self, name: str) -> tuple[DriveResource, str, int]:
        marker_query = (
            f"appProperties has {{ key='{SHEET_MARKER_KEY}' and "
            f"value='{SHEET_MARKER_VALUE}' }}"
        )
        resource = self._find_resource(
            self.parent_folder_id,
            name,
            GOOGLE_SHEET_MIME,
            extra_query=marker_query,
        )
        created = resource is None
        if resource is None:
            resource = self._create_resource(
                name,
                GOOGLE_SHEET_MIME,
                self.parent_folder_id,
                app_properties={SHEET_MARKER_KEY: SHEET_MARKER_VALUE},
            )

        sheet_id, sheet_title, row_count = self._first_sheet(resource.id)
        if created and sheet_title != "Tin nhắn":
            self._rename_sheet(resource.id, sheet_id, "Tin nhắn")
            sheet_title = "Tin nhắn"
        self._ensure_sheet_header(resource.id, sheet_id, sheet_title)
        return resource, sheet_title, row_count

    def _ensure_image_folder(self, name: str) -> DriveResource:
        resource = self._find_resource(
            self.parent_folder_id,
            name,
            GOOGLE_FOLDER_MIME,
        )
        if resource is not None:
            return resource
        return self._create_resource(name, GOOGLE_FOLDER_MIME, self.parent_folder_id)

    def _find_resource(
        self,
        parent_id: str,
        name: str,
        mime_type: str,
        *,
        extra_query: str | None = None,
    ) -> DriveResource | None:
        query = (
            f"'{self._query_literal(parent_id)}' in parents and trashed = false "
            f"and name = '{self._query_literal(name)}' "
            f"and mimeType = '{self._query_literal(mime_type)}'"
        )
        if extra_query:
            query += f" and {extra_query}"
        payload = self._request_json(
            "GET",
            f"{DRIVE_API}/files",
            params={
                "q": query,
                "spaces": "drive",
                "pageSize": 10,
                "orderBy": "createdTime desc",
                "includeItemsFromAllDrives": "true",
                "supportsAllDrives": "true",
                "fields": "files(id,name,mimeType,webViewLink)",
            },
        )
        files = payload.get("files") or []
        if not files:
            return None
        return self._resource(files[0])

    def _create_resource(
        self,
        name: str,
        mime_type: str,
        parent_id: str,
        *,
        app_properties: dict[str, str] | None = None,
    ) -> DriveResource:
        metadata: dict[str, Any] = {
            "name": name,
            "mimeType": mime_type,
            "parents": [parent_id],
        }
        if app_properties:
            metadata["appProperties"] = app_properties
        payload = self._request_json(
            "POST",
            f"{DRIVE_API}/files",
            expected=(200, 201),
            params={
                "supportsAllDrives": "true",
                "fields": "id,name,mimeType,webViewLink",
            },
            json=metadata,
        )
        resource_id = str(payload.get("id") or "")
        if not resource_id:
            raise GoogleDriveOutputError("Google Drive không trả id tài nguyên mới.")
        verified = self._request_json(
            "GET",
            f"{DRIVE_API}/files/{quote(resource_id, safe='')}",
            params={
                "supportsAllDrives": "true",
                "fields": "id,name,mimeType,webViewLink",
            },
        )
        return self._resource(verified)

    def _first_sheet(self, spreadsheet_id: str) -> tuple[int, str, int]:
        payload = self._request_json(
            "GET",
            f"{SHEETS_API}/spreadsheets/{quote(spreadsheet_id, safe='')}",
            params={
                "fields": "sheets(properties(sheetId,title,gridProperties(rowCount)))"
            },
        )
        sheets = payload.get("sheets") or []
        if not sheets:
            raise GoogleDriveOutputError("Google Sheet không có tab để ghi dữ liệu.")
        properties = sheets[0].get("properties") or {}
        try:
            grid_properties = properties.get("gridProperties") or {}
            return (
                int(properties["sheetId"]),
                str(properties["title"]),
                max(1, int(grid_properties.get("rowCount") or 1000)),
            )
        except (KeyError, TypeError, ValueError) as exc:
            raise GoogleDriveOutputError("Metadata tab Google Sheet không hợp lệ.") from exc

    def _rename_sheet(
        self, spreadsheet_id: str, sheet_id: int, new_title: str
    ) -> None:
        self._sheets_batch_update(
            spreadsheet_id,
            [
                {
                    "updateSheetProperties": {
                        "properties": {"sheetId": sheet_id, "title": new_title},
                        "fields": "title",
                    }
                }
            ],
        )

    def _ensure_sheet_header(
        self,
        spreadsheet_id: str,
        sheet_id: int,
        sheet_title: str,
    ) -> None:
        header_range = f"{self._a1_sheet(sheet_title)}!A1:J1"
        values = self._get_values(spreadsheet_id, header_range)
        if values:
            existing = [str(value) for value in values[0]]
            if existing == list(SHEET_HEADERS):
                return
            if existing == list(LEGACY_SHEET_HEADERS):
                self._request_json(
                    "PUT",
                    f"{SHEETS_API}/spreadsheets/"
                    f"{quote(spreadsheet_id, safe='')}/values/"
                    f"{quote(f'{self._a1_sheet(sheet_title)}!J1', safe='')}",
                    params={"valueInputOption": "RAW"},
                    json={"majorDimension": "ROWS", "values": [["Chi nhánh"]]},
                )
                self._format_migrated_branch_column(spreadsheet_id, sheet_id)
                return
            if existing != list(SHEET_HEADERS):
                raise GoogleDriveOutputError(
                    f"Sheet {spreadsheet_id} đã có header khác cấu trúc crawler; "
                    "tool dừng để không ghi đè dữ liệu."
                )

        self._request_json(
            "PUT",
            f"{SHEETS_API}/spreadsheets/{quote(spreadsheet_id, safe='')}/values/"
            f"{quote(header_range, safe='')}",
            params={"valueInputOption": "RAW"},
            json={"majorDimension": "ROWS", "values": [list(SHEET_HEADERS)]},
        )
        self._sheets_batch_update(
            spreadsheet_id,
            [
                {
                    "updateSheetProperties": {
                        "properties": {
                            "sheetId": sheet_id,
                            "gridProperties": {"frozenRowCount": 1},
                        },
                        "fields": "gridProperties.frozenRowCount",
                    }
                },
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 0,
                            "endRowIndex": 1,
                            "startColumnIndex": 0,
                            "endColumnIndex": len(SHEET_HEADERS),
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColorStyle": {
                                    "rgbColor": {
                                        "red": 0.035,
                                        "green": 0.408,
                                        "blue": 1.0,
                                    }
                                },
                                "textFormat": {
                                    "bold": True,
                                    "foregroundColorStyle": {
                                        "rgbColor": {
                                            "red": 1.0,
                                            "green": 1.0,
                                            "blue": 1.0,
                                        }
                                    },
                                },
                            }
                        },
                        "fields": "userEnteredFormat(backgroundColorStyle,textFormat)",
                    }
                },
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 1,
                            "startColumnIndex": 0,
                            "endColumnIndex": len(SHEET_HEADERS),
                        },
                        "cell": {
                            "userEnteredFormat": {"verticalAlignment": "TOP"}
                        },
                        "fields": "userEnteredFormat.verticalAlignment",
                    }
                },
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 1,
                            "startColumnIndex": 7,
                            "endColumnIndex": 8,
                        },
                        "cell": {
                            "userEnteredFormat": {"wrapStrategy": "WRAP"}
                        },
                        "fields": "userEnteredFormat.wrapStrategy",
                    }
                },
                *self._column_width_requests(sheet_id),
            ],
        )

    def _append_unique_text_rows(
        self,
        spreadsheet_id: str,
        sheet_title: str,
        row_count: int,
        rows: list[list[Any]],
    ) -> int:
        existing_keys: set[tuple[str, str]] = set()
        chunk_size = 20_000
        for start_row in range(2, row_count + 1, chunk_size):
            end_row = min(row_count, start_row + chunk_size - 1)
            key_range = (
                f"{self._a1_sheet(sheet_title)}!B{start_row}:C{end_row}"
            )
            existing_values = self._get_values(spreadsheet_id, key_range)
            existing_keys.update(
                (str(row[0]), str(row[1]))
                for row in existing_values
                if len(row) >= 2
            )
        pending = [
            row for row in rows if (str(row[1]), str(row[2])) not in existing_keys
        ]
        if not pending:
            return 0

        append_range = f"{self._a1_sheet(sheet_title)}!A:J"
        self._request_json(
            "POST",
            f"{SHEETS_API}/spreadsheets/{quote(spreadsheet_id, safe='')}/values/"
            f"{quote(append_range, safe='')}:append",
            params={
                "valueInputOption": "RAW",
                "insertDataOption": "INSERT_ROWS",
            },
            json={"majorDimension": "ROWS", "values": pending},
        )
        return len(pending)

    def _get_values(self, spreadsheet_id: str, range_name: str) -> list[list[Any]]:
        payload = self._request_json(
            "GET",
            f"{SHEETS_API}/spreadsheets/{quote(spreadsheet_id, safe='')}/values/"
            f"{quote(range_name, safe='')}",
            params={"majorDimension": "ROWS"},
        )
        values = payload.get("values") or []
        return [row for row in values if isinstance(row, list)]

    def _sheets_batch_update(
        self, spreadsheet_id: str, requests: list[dict[str, Any]]
    ) -> None:
        if not requests or any(len(request) != 1 for request in requests):
            raise GoogleDriveOutputError(
                "Mỗi Sheets batch request phải có đúng một loại thao tác."
            )
        self._request_json(
            "POST",
            f"{SHEETS_API}/spreadsheets/{quote(spreadsheet_id, safe='')}:batchUpdate",
            json={"requests": requests},
        )

    def _format_migrated_branch_column(
        self, spreadsheet_id: str, sheet_id: int
    ) -> None:
        self._sheets_batch_update(
            spreadsheet_id,
            [
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 0,
                            "endRowIndex": 1,
                            "startColumnIndex": 9,
                            "endColumnIndex": 10,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "backgroundColorStyle": {
                                    "rgbColor": {
                                        "red": 0.035,
                                        "green": 0.408,
                                        "blue": 1.0,
                                    }
                                },
                                "textFormat": {
                                    "bold": True,
                                    "foregroundColorStyle": {
                                        "rgbColor": {
                                            "red": 1.0,
                                            "green": 1.0,
                                            "blue": 1.0,
                                        }
                                    },
                                },
                            }
                        },
                        "fields": (
                            "userEnteredFormat(backgroundColorStyle,textFormat)"
                        ),
                    }
                },
                self._column_width_requests(sheet_id)[9],
            ],
        )

    def _upload_unique_image(
        self,
        folder_id: str,
        *,
        group_name: str,
        message_id: str,
        media: MediaAsset,
        media_path: Path,
    ) -> bool:
        digest = media.sha256 or hashlib.sha256(media_path.read_bytes()).hexdigest()
        source_key = hashlib.sha256(
            f"{group_name}\0{message_id}\0{digest}".encode("utf-8")
        ).hexdigest()
        filename = f"{safe_slug(group_name)}-{media_path.name}"
        marker_query = (
            f"appProperties has {{ key='{IMAGE_MARKER_KEY}' and "
            f"value='{source_key}' }}"
        )
        existing = self._find_resource(
            folder_id,
            filename,
            media.mime_type,
            extra_query=marker_query,
        )
        if existing is not None:
            return False

        metadata = {
            "name": filename,
            "parents": [folder_id],
            "appProperties": {IMAGE_MARKER_KEY: source_key},
        }
        boundary = f"codex-{uuid.uuid4().hex}"
        metadata_bytes = json.dumps(metadata, ensure_ascii=False).encode("utf-8")
        file_bytes = media_path.read_bytes()
        body = b"\r\n".join(
            [
                f"--{boundary}".encode(),
                b"Content-Type: application/json; charset=UTF-8",
                b"",
                metadata_bytes,
                f"--{boundary}".encode(),
                f"Content-Type: {media.mime_type}".encode(),
                b"",
                file_bytes,
                f"--{boundary}--".encode(),
                b"",
            ]
        )
        self._request_json(
            "POST",
            f"{DRIVE_UPLOAD_API}/files",
            expected=(200, 201),
            params={
                "uploadType": "multipart",
                "supportsAllDrives": "true",
                "fields": "id,name",
            },
            headers={"Content-Type": f"multipart/related; boundary={boundary}"},
            data=body,
        )
        return True

    def _request_json(
        self,
        method: str,
        url: str,
        *,
        expected: tuple[int, ...] = (200,),
        **kwargs: Any,
    ) -> dict[str, Any]:
        try:
            response = self.session.request(
                method, url, timeout=120, **kwargs
            )
        except Exception as exc:
            raise GoogleDriveOutputError(
                f"Không kết nối được Google API: {exc}"
            ) from exc
        try:
            payload = response.json()
        except Exception:
            payload = {}
        if response.status_code not in expected:
            error = payload.get("error") if isinstance(payload, dict) else None
            if isinstance(error, dict):
                detail = str(error.get("message") or "")
            else:
                detail = str(error or "")
            if not detail:
                detail = str(getattr(response, "text", ""))[:500]
            raise GoogleDriveOutputError(
                f"Google API trả HTTP {response.status_code}: "
                f"{detail or 'không có chi tiết lỗi'}"
            )
        if not isinstance(payload, dict):
            raise GoogleDriveOutputError("Google API trả dữ liệu không hợp lệ.")
        return payload

    @staticmethod
    def _resource(payload: dict[str, Any]) -> DriveResource:
        resource_id = str(payload.get("id") or "")
        name = str(payload.get("name") or "")
        url = str(payload.get("webViewLink") or "")
        if not resource_id or not name or not url:
            raise GoogleDriveOutputError(
                "Google Drive không trả đủ id, tên và link của tài nguyên."
            )
        return DriveResource(id=resource_id, name=name, url=url)

    @staticmethod
    def _query_literal(value: str) -> str:
        return value.replace("\\", "\\\\").replace("'", "\\'")

    @staticmethod
    def _a1_sheet(title: str) -> str:
        return "'" + title.replace("'", "''") + "'"

    @staticmethod
    def _column_width_requests(sheet_id: int) -> list[dict[str, Any]]:
        widths = (96, 180, 220, 60, 160, 90, 110, 420, 90, 200)
        return [
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": index,
                        "endIndex": index + 1,
                    },
                    "properties": {"pixelSize": pixel_size},
                    "fields": "pixelSize",
                }
            }
            for index, pixel_size in enumerate(widths)
        ]

    @staticmethod
    def _configured_path(value: str, base_dir: Path) -> Path:
        path = Path(value).expanduser()
        return path if path.is_absolute() else base_dir / path

    @staticmethod
    def _oauth_redirect_port() -> int:
        raw = os.environ.get("GOOGLE_OAUTH_REDIRECT_PORT", "8766").strip()
        try:
            port = int(raw)
        except ValueError as exc:
            raise GoogleDriveOutputError(
                "GOOGLE_OAUTH_REDIRECT_PORT phải là số nguyên."
            ) from exc
        if not 1 <= port <= 65535:
            raise GoogleDriveOutputError(
                "GOOGLE_OAUTH_REDIRECT_PORT phải nằm trong khoảng 1-65535."
            )
        return port
