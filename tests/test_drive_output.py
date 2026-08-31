from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from zalo_order_crawler.drive_output import (
    DriveResource,
    GoogleDriveOutputError,
    GoogleDriveOutputPublisher,
    OCR_HEADERS,
)
from zalo_order_crawler.models import (
    CleanMessage,
    ImageOcrResult,
    MediaAsset,
    OcrLineItem,
    OrderDecision,
)


class FakeResponse:
    status_code = 200
    text = ""

    def __init__(self, payload: dict[str, Any]) -> None:
        self.payload = payload

    def json(self) -> dict[str, Any]:
        return self.payload


class FakeSession:
    def __init__(self, payload: dict[str, Any]) -> None:
        self.payload = payload

    def request(self, *_: Any, **__: Any) -> FakeResponse:
        return FakeResponse(self.payload)


class RecordingPublisher(GoogleDriveOutputPublisher):
    def __init__(self) -> None:
        super().__init__(session=object(), parent_folder_id="folder123")
        self.sheet_name = ""
        self.folder_name = ""
        self.rows: list[list[Any]] = []
        self.uploaded: list[tuple[str, str, str]] = []
        self.ocr_workbook_name = ""
        self.ocr_rows: list[list[Any]] = []

    def _ensure_daily_sheet(self, name: str) -> tuple[DriveResource, str, int]:
        self.sheet_name = name
        return (
            DriveResource(
                id="sheet-id",
                name=name,
                url="https://docs.google.com/spreadsheets/d/sheet-id/edit",
            ),
            "Tin nhắn",
            1000,
        )

    def _append_unique_text_rows(
        self,
        spreadsheet_id: str,
        sheet_title: str,
        row_count: int,
        rows: list[list[Any]],
    ) -> int:
        assert spreadsheet_id == "sheet-id"
        assert sheet_title == "Tin nhắn"
        assert row_count == 1000
        self.rows = rows
        return len(rows)

    def _ensure_image_folder(self, name: str) -> DriveResource:
        self.folder_name = name
        return DriveResource(
            id="image-folder-id",
            name=name,
            url="https://drive.google.com/drive/folders/image-folder-id",
        )

    def _upload_unique_image(
        self,
        folder_id: str,
        *,
        group_name: str,
        message_id: str,
        media: MediaAsset,
        media_path: Path,
    ) -> bool:
        assert folder_id == "image-folder-id"
        self.uploaded.append((group_name, message_id, media_path.name))
        return True

    def _ensure_ocr_workbook(self, name: str) -> DriveResource:
        self.ocr_workbook_name = name
        return DriveResource(
            id="ocr-workbook-id",
            name=name,
            url="https://drive.google.com/file/d/ocr-workbook-id/view",
        )

    def _append_unique_ocr_rows(self, file_id: str, rows: list[list[Any]]) -> int:
        assert file_id == "ocr-workbook-id"
        self.ocr_rows = rows
        return len(rows)


def test_publish_splits_text_and_message_images_by_selected_date(tmp_path: Path) -> None:
    assets = tmp_path / "assets"
    assets.mkdir()
    image = assets / "order.jpg"
    image.write_bytes(b"jpeg")
    thumbnail = assets / "thumbnail.jpg"
    thumbnail.write_bytes(b"thumbnail")
    messages = [
        CleanMessage(
            message_id="text-1",
            sequence=1,
            sender="Lan",
            timestamp_text="09:15",
            content="Chốt 2 áo",
            direction="incoming",
            message_type="text",
        ),
        CleanMessage(
            message_id="image-1",
            sequence=2,
            content="[Hình ảnh]",
            message_type="image",
            media=[
                MediaAsset(
                    path="assets/order.jpg",
                    mime_type="image/jpeg",
                    role="message_image",
                    sha256="abc",
                ),
                MediaAsset(
                    path="assets/thumbnail.jpg",
                    mime_type="image/jpeg",
                    role="link_thumbnail",
                    sha256="def",
                ),
            ],
        ),
    ]
    publisher = RecordingPublisher()

    result = publisher.publish(
        run_dir=tmp_path,
        group_name="Nhóm A",
        target_date=date(2026, 8, 30),
        messages=messages,
    )

    assert publisher.sheet_name == "30-08-2026"
    assert publisher.folder_name == "30-08-2026_image"
    assert publisher.rows == [
        [
            "30-08-2026",
            "Nhóm A",
            "text-1",
            1,
            "Lan",
            "09:15",
            "incoming",
            "Chốt 2 áo",
            "text",
            "",
        ]
    ]
    assert publisher.uploaded == [("Nhóm A", "image-1", "order.jpg")]
    assert result["sheet"]["rows_added"] == 1
    assert result["image_folder"]["images_uploaded"] == 1


def test_publish_uploads_ocr_workbook_when_rows_present(tmp_path: Path) -> None:
    publisher = RecordingPublisher()

    result = publisher.publish(
        run_dir=tmp_path,
        group_name="Nhóm A",
        target_date=date(2026, 8, 30),
        messages=[],
        ocr_results=[
            ImageOcrResult(
                message_id="m1",
                media_path="assets/m1.jpg",
                applicable=True,
                items=[
                    OcrLineItem(
                        customer_code="S6",
                        customer_name="Quán Rau",
                        product_name="Rau muống",
                        unit="kg",
                        quantity=2,
                    )
                ],
            )
        ],
    )

    assert publisher.ocr_workbook_name == "30-08-2026_OCR.xlsx"
    assert publisher.ocr_rows == [
        ["30-08-2026", "Nhóm A", "m1", "S6", "Quán Rau", "Rau muống", "kg", 2]
    ]
    assert result["ocr_workbook"]["rows_added"] == 1


def test_publish_skips_ocr_workbook_when_no_applicable_results(tmp_path: Path) -> None:
    publisher = RecordingPublisher()

    result = publisher.publish(
        run_dir=tmp_path,
        group_name="Nhóm A",
        target_date=date(2026, 8, 30),
        messages=[],
        ocr_results=[
            ImageOcrResult(
                message_id="m1", media_path="assets/m1.jpg", applicable=False
            )
        ],
    )

    assert publisher.ocr_workbook_name == ""
    assert "ocr_workbook" not in result


def test_ocr_rows_skip_non_applicable_and_blank_product_name() -> None:
    ocr_results = [
        ImageOcrResult(
            message_id="m1",
            media_path="assets/m1.jpg",
            applicable=True,
            items=[
                OcrLineItem(product_name="Rau muống", unit="kg", quantity=2),
                OcrLineItem(product_name="   "),
            ],
        ),
        ImageOcrResult(message_id="m2", media_path="assets/m2.jpg", applicable=False),
    ]

    rows = GoogleDriveOutputPublisher._ocr_rows("Nhóm A", date(2026, 8, 30), ocr_results)

    assert rows == [["30-08-2026", "Nhóm A", "m1", "", "", "Rau muống", "kg", 2]]


def test_append_unique_ocr_rows_skips_existing_group_message_product_key() -> None:
    publisher = GoogleDriveOutputPublisher(object(), "folder123")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(list(OCR_HEADERS))
    sheet.append(
        ["30-08-2026", "Nhóm A", "m1", "S6", "Quán Rau", "Rau muống", "kg", 2]
    )
    buffer = BytesIO()
    workbook.save(buffer)
    publisher._download_file = lambda file_id: buffer.getvalue()  # type: ignore[method-assign]
    saved: dict[str, Any] = {}

    def fake_update(file_id: str, mime_type: str, data: bytes) -> None:
        saved["file_id"] = file_id
        saved["workbook"] = openpyxl.load_workbook(BytesIO(data))

    publisher._update_file_content = fake_update  # type: ignore[method-assign]
    rows = [
        ["30-08-2026", "Nhóm A", "m1", "S6", "Quán Rau", "Rau muống", "kg", 2],
        ["30-08-2026", "Nhóm A", "m1", "S6", "Quán Rau", "Cải ngọt", "kg", 1],
    ]

    added = publisher._append_unique_ocr_rows("workbook-id", rows)

    assert added == 1
    assert saved["file_id"] == "workbook-id"
    saved_rows = list(saved["workbook"].active.iter_rows(values_only=True))
    assert saved_rows[-1] == (
        "30-08-2026",
        "Nhóm A",
        "m1",
        "S6",
        "Quán Rau",
        "Cải ngọt",
        "kg",
        1,
    )
    assert len(saved_rows) == 3


def test_append_unique_text_rows_skips_existing_group_message_key() -> None:
    publisher = GoogleDriveOutputPublisher(object(), "folder123")
    requests: list[dict[str, Any]] = []
    publisher._get_values = lambda *_: [["Nhóm A", "m1"]]  # type: ignore[method-assign]
    publisher._request_json = (  # type: ignore[method-assign]
        lambda method, url, **kwargs: requests.append(
            {"method": method, "url": url, **kwargs}
        )
        or {}
    )
    rows = [
        ["30-08-2026", "Nhóm A", "m1", 1, "", "", "incoming", "Cũ", "text", ""],
        ["30-08-2026", "Nhóm B", "m2", 2, "", "", "incoming", "Mới", "text", ""],
    ]

    added = publisher._append_unique_text_rows(
        "sheet-id", "Tin nhắn", 1000, rows
    )

    assert added == 1
    assert requests[0]["json"]["values"] == [rows[1]]


def test_media_path_cannot_escape_run_directory(tmp_path: Path) -> None:
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    outside = tmp_path / "outside.jpg"
    outside.write_bytes(b"private")

    with pytest.raises(GoogleDriveOutputError, match="ngoài thư mục"):
        GoogleDriveOutputPublisher._resolve_media_path(run_dir, "../outside.jpg")


def test_text_rows_include_branch_only_for_order_decisions() -> None:
    messages = [
        CleanMessage(message_id="m1", sequence=1, content="S6 thêm 2 kg rau"),
        CleanMessage(message_id="m2", sequence=2, content="Dạ"),
    ]
    decisions = [
        OrderDecision(
            message_id="m1",
            is_order=True,
            confidence=0.98,
            data_confidence=0.95,
            reason="Đơn bổ sung",
            branch_name="Chi nhánh Phạm Văn Đồng",
            products=["rau"],
            quantities=["2 kg"],
        ),
        OrderDecision(
            message_id="m2",
            is_order=False,
            confidence=0.99,
            data_confidence=0,
            reason="Xã giao",
            branch_name="Chi nhánh Tân Phú",
        ),
    ]

    rows = GoogleDriveOutputPublisher._text_rows(
        "Rau SMO", date(2026, 8, 30), messages, decisions
    )

    assert rows[0][-1] == "Chi nhánh Phạm Văn Đồng"
    assert rows[1][-1] == ""


def test_parse_branch_mappings_preserves_aliases_and_rejects_conflicts() -> None:
    rows = [
        ["S6", "Chi nhánh Phạm Văn Đồng"],
        ["Tân Phú", "Chi nhánh Tân Phú"],
    ]

    assert GoogleDriveOutputPublisher._parse_branch_mappings(rows) == {
        "S6": "Chi nhánh Phạm Văn Đồng",
        "Tân Phú": "Chi nhánh Tân Phú",
    }

    with pytest.raises(GoogleDriveOutputError, match="nhiều chi nhánh"):
        GoogleDriveOutputPublisher._parse_branch_mappings(
            [["S6", "Chi nhánh A"], [" s6 ", "Chi nhánh B"]]
        )


def test_verify_destination_requires_permission_to_add_files() -> None:
    publisher = GoogleDriveOutputPublisher(
        FakeSession(
            {
                "id": "folder123",
                "name": "Output",
                "mimeType": "application/vnd.google-apps.folder",
                "webViewLink": "https://drive.google.com/drive/folders/folder123",
                "trashed": False,
                "capabilities": {"canAddChildren": False},
            }
        ),
        "folder123",
    )

    with pytest.raises(GoogleDriveOutputError, match="không có quyền"):
        publisher.verify_destination()


def test_web_oauth_redirect_port_validation(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("GOOGLE_OAUTH_REDIRECT_PORT", "8766")
    assert GoogleDriveOutputPublisher._oauth_redirect_port() == 8766

    monkeypatch.setenv("GOOGLE_OAUTH_REDIRECT_PORT", "invalid")
    with pytest.raises(GoogleDriveOutputError, match="số nguyên"):
        GoogleDriveOutputPublisher._oauth_redirect_port()
